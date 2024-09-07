import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import os
import sys  # 用于在用户选择后退出程序

# 文件路径
file_A = 'A.xlsx'
file_B = 'B.xlsx'
file_TEMP = 'TEMP.xlsx'
result_new_path = 'result_new.xlsx'

# 读取Excel文件
df_A = pd.read_excel(file_A)
df_B = pd.read_excel(file_B)

# 只检测A.xlsx中使用到的列（'数量', '采购单价 含税', '代码'）
required_columns_A = ['数量', '采购单价 含税', '代码']
missing_data_A = df_A[required_columns_A].isnull()  # 判断这些列中哪些位置是缺失数据
missing_positions_A = missing_data_A[missing_data_A == True].stack().index.tolist()  # 找到缺失数据的位置

# 只检测B.xlsx中使用到的列（'备注', '价税合计'）
required_columns_B = ['备注', '价税合计']
missing_data_B = df_B[required_columns_B].isnull()  # 判断这些列中哪些位置是缺失数据
missing_positions_B = missing_data_B[missing_data_B == True].stack().index.tolist()  # 找到缺失数据的位置

# 处理A.xlsx中的缺失数据
if missing_positions_A:
    print("A.xlsx 文件中存在缺失数据，缺失的位置如下：")
    for row, col in missing_positions_A:
        print(f"第 {row + 2} 行，列 '{col}' 缺失数据。")
    
    user_input = input("输入 'continue' 无视缺失数据继续运行，或按回车键退出程序: ")
    if user_input.lower() != 'continue':
        sys.exit("程序已退出，请修复数据后重试。")

# 处理B.xlsx中的缺失数据
if missing_positions_B:
    print("B.xlsx 文件中存在缺失数据，缺失的位置如下：")
    for row, col in missing_positions_B:
        print(f"第 {row + 2} 行，列 '{col}' 缺失数据。")
    
    user_input = input("输入 'continue' 无视缺失数据继续运行，或按回车键退出程序: ")
    if user_input.lower() != 'continue':
        sys.exit("程序已退出，请修复数据后重试。")

# 调试: 检查初始数据框
print("DataFrame A:")
print(df_A.head())
print("\nDataFrame B:")
print(df_B.head())

# 将相关列转换为数值类型，遇到错误时强制转换为NaN
df_A['数量'] = pd.to_numeric(df_A['数量'], errors='coerce')
df_A['采购单价 含税'] = pd.to_numeric(df_A['采购单价 含税'], errors='coerce')

# 将NaN值替换为0，以便进行乘法运算
df_A['数量'].fillna(0, inplace=True)
df_A['采购单价 含税'].fillna(0, inplace=True)

# 计算 '总价' 并按 '代码' 进行汇总
df_A['总价'] = df_A['数量'] * df_A['采购单价 含税']
df_A_summary = df_A.groupby('代码')['总价'].sum().reset_index()

# 调试: 检查汇总后的A表
print("\nSummary of A:")
print(df_A_summary.head())

# 保存 TEMP.xlsx
wb_temp = Workbook()
ws_temp = wb_temp.active
ws_temp.title = 'TEMP'
ws_temp.append(['代码', '总价'])
for index, row in df_A_summary.iterrows():
    ws_temp.append(row.tolist())

wb_temp.save(file_TEMP)

# 按 '备注' 汇总 B 表
df_B_summary = df_B.groupby('备注')['价税合计'].sum().reset_index()

# 创建结果数据框
result = df_B_summary.copy()
result['代码'] = ''
result['总价'] = 0.0
result['差值'] = 0.0

# 保存初始的 result_new.xlsx
wb_result = Workbook()
ws_result = wb_result.active
ws_result.title = 'Result'
ws_result.append(['备注', '价税合计', '代码', '总价', '差值'])

for index, row in result.iterrows():
    ws_result.append([
        row['备注'],
        row['价税合计'],
        '',
        0.0,
        0.0
    ])

wb_result.save(result_new_path)

# 读取 TEMP.xlsx
df_TEMP = pd.read_excel(file_TEMP)

# 确保 TEMP 列是字符串类型并删除多余的空格
df_TEMP['代码'] = df_TEMP['代码'].astype(str).str.strip()

# 重新读取 result_new.xlsx 以合并 TEMP 数据
wb_result_new = load_workbook(filename=result_new_path)
ws_result_new = wb_result_new.active

# 填充 result 数据框中的 '代码' 和 '总价' 列
for temp_index, temp_row in df_TEMP.iterrows():
    code = temp_row['代码']
    total_price = temp_row['总价']
    for result_index, result_row in enumerate(ws_result_new.iter_rows(min_row=2, values_only=True), start=2):
        if str(result_row[0]).strip() == code:
            ws_result_new.cell(row=result_index, column=3, value=code)
            ws_result_new.cell(row=result_index, column=4, value=total_price)

# 计算 '差值'
for result_index, result_row in enumerate(ws_result_new.iter_rows(min_row=2, values_only=True), start=2):
    price_tax = result_row[1]
    total_price = result_row[3]
    diff = total_price - price_tax
    ws_result_new.cell(row=result_index, column=5, value=round(diff, 2))

wb_result_new.save(result_new_path)

# 在 result_new.xlsx 中查找 '差值' 为0的行，标记 A.xlsx 中的相应代码
wb_A = load_workbook(filename=file_A)
ws_A = wb_A.active

# 查找 result_new.xlsx 中 '差值' 为 0 的代码
codes_to_highlight = set()
for row in ws_result_new.iter_rows(min_row=2, values_only=True):
    if row[4] == 0:
        code = str(row[2])
        codes_to_highlight.add(code)

# 标记 A.xlsx 中匹配的代码行
for row in ws_A.iter_rows(min_row=2, max_row=ws_A.max_row):
    code_a = str(row[df_A.columns.get_loc('代码')].value).strip()
    if code_a in codes_to_highlight:
        for cell in row:
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# 保存更新后的 A.xlsx 文件
highlighted_file_A = 'A_highlighted.xlsx'
wb_A.save(highlighted_file_A)

# 返回更新后的文件路径
highlighted_file_A, result_new_path

# 在程序最后等待用户按Enter键退出
input("程序运行完毕。按回车键退出...")
